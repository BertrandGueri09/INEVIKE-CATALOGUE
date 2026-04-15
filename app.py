import streamlit as st
import pandas as pd
import json
import os
import uuid
from copy import deepcopy
from datetime import date, timedelta
from io import BytesIO

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


st.set_page_config(
    page_title="INEVOKE — Outil de Devis",
    page_icon="☀️",
    layout="wide",
    initial_sidebar_state="expanded",
)

DB_FILE = "equipements_db.json"
SETTINGS_FILE = "catalogue_settings.json"

DEFAULT_FOURNISSEURS = ["SOGELUX", "DEYE", "HONLE", "ECS", "AUTRES"]
DEFAULT_CATEGORIES = [
    "Panneaux",
    "Onduleur",
    "Batteries",
    "Structure",
    "Protections DC",
    "Protections AC",
    "Protections Batterie",
    "Câbles",
    "Accessoires",
    "Main d'œuvre",
]
UNITES = ["pcs", "lot", "m", "kit", "barre", "boite", "rouleaux", "ml", "unité"]

NAVY = "#0A2540"
ORANGE = "#F5A623"


# ──────────────────────────────────────────────────────────────────────────────
# Session state
# ──────────────────────────────────────────────────────────────────────────────
def init_session_state():
    defaults = {
        "lignes_devis": [],
        "devis_pdf_bytes": None,
        "devis_pdf_name": None,
        "devis_total_ttc": None,
        "flash_success": "",
        "flash_warning": "",
        "flash_info": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_session_state()


# ──────────────────────────────────────────────────────────────────────────────
# Paramètres
# ──────────────────────────────────────────────────────────────────────────────
def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        return {
            "fournisseurs": DEFAULT_FOURNISSEURS.copy(),
            "categories": DEFAULT_CATEGORIES.copy(),
        }
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        if "fournisseurs" not in data or not isinstance(data["fournisseurs"], list):
            data["fournisseurs"] = DEFAULT_FOURNISSEURS.copy()

        if "categories" not in data or not isinstance(data["categories"], list):
            data["categories"] = DEFAULT_CATEGORIES.copy()

        return data
    except Exception:
        return {
            "fournisseurs": DEFAULT_FOURNISSEURS.copy(),
            "categories": DEFAULT_CATEGORIES.copy(),
        }


def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


settings = load_settings()
FOURNISSEURS = list(settings.get("fournisseurs", DEFAULT_FOURNISSEURS.copy()))
CATEGORIES = DEFAULT_CATEGORIES.copy()
for cat in settings.get("categories", []):
    if cat and cat not in CATEGORIES:
        CATEGORIES.append(cat)


def persist_categories(categories):
    uniq = []
    for c in categories:
        c = str(c).strip()
        if c and c not in uniq:
            uniq.append(c)
    settings["categories"] = uniq
    save_settings(settings)


def persist_fournisseurs(fournisseurs):
    uniq = []
    for f in fournisseurs:
        f = str(f).strip().upper()
        if f and f not in uniq:
            uniq.append(f)
    settings["fournisseurs"] = uniq
    save_settings(settings)


# ──────────────────────────────────────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(
    f"""
<style>
.main-header{{background:{NAVY};color:#fff;padding:18px 24px;border-radius:10px;
  margin-bottom:1.5rem;display:flex;align-items:center;gap:16px}}
.main-header h1{{margin:0;font-size:22px;font-weight:600;color:#fff}}
.main-header p{{margin:0;font-size:13px;color:#9ab3c8}}
.section-title{{font-size:15px;font-weight:600;color:{NAVY};
  border-left:4px solid {ORANGE};padding-left:10px;margin:1.2rem 0 .8rem}}
.kpi-card{{background:#f4f6f8;border-radius:8px;padding:14px 16px;text-align:center}}
.kpi-val{{font-size:24px;font-weight:700;color:{NAVY}}}
.kpi-lbl{{font-size:11px;color:#666;margin-top:2px}}
.success-box{{background:#e8f5e9;border:1px solid #4caf50;border-radius:8px;
  padding:10px 14px;color:#1b5e20;font-size:13px;margin:.5rem 0}}
.warn-box{{background:#fff8e1;border:1px solid #ffc107;border-radius:8px;
  padding:10px 14px;color:#5d4037;font-size:13px;margin:.5rem 0}}
.info-box{{background:#e3f2fd;border:1px solid #2196f3;border-radius:8px;
  padding:10px 14px;color:#0d47a1;font-size:13px;margin:.5rem 0}}
.stButton>button{{border-radius:6px;font-weight:500}}
</style>
""",
    unsafe_allow_html=True,
)


# ──────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ──────────────────────────────────────────────────────────────────────────────
def gen_id():
    return "eq_" + uuid.uuid4().hex[:8]


def safe_float(v):
    try:
        if pd.isna(v):
            return None
        fv = float(v)
        return fv if fv > 0 else None
    except Exception:
        return None


def fmt_fcfa(v):
    return f"{v:,.0f}".replace(",", "  ") if v and v > 0 else "—"


def normalize_record(r, fournisseurs=None):
    fournisseurs = fournisseurs or FOURNISSEURS
    nr = {
        "id": r.get("id") or gen_id(),
        "categorie": str(r.get("categorie", "Accessoires")).strip() or "Accessoires",
        "designation": str(r.get("designation", "")).strip(),
        "unite": str(r.get("unite", "pcs")).strip() or "pcs",
        "note": str(r.get("note", "")).strip(),
    }
    for fn in fournisseurs:
        nr[fn] = safe_float(r.get(fn))
    return nr


def ensure_all_supplier_keys(data, fournisseurs=None):
    fournisseurs = fournisseurs or FOURNISSEURS
    return [normalize_record(r, fournisseurs=fournisseurs) for r in data]


def show_flash_messages():
    if st.session_state.flash_success:
        st.success(st.session_state.flash_success)
        st.session_state.flash_success = ""
    if st.session_state.flash_warning:
        st.warning(st.session_state.flash_warning)
        st.session_state.flash_warning = ""
    if st.session_state.flash_info:
        st.info(st.session_state.flash_info)
        st.session_state.flash_info = ""


# ──────────────────────────────────────────────────────────────────────────────
# Base de données
# ──────────────────────────────────────────────────────────────────────────────
def load_db():
    if not os.path.exists(DB_FILE):
        return []
    try:
        with open(DB_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return ensure_all_supplier_keys(data, fournisseurs=FOURNISSEURS)
    except Exception:
        return []


def save_db(data):
    clean = ensure_all_supplier_keys(data, fournisseurs=FOURNISSEURS)
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)


def db_to_df(db):
    base_cols = ["id", "categorie", "designation", "unite", "note"] + FOURNISSEURS
    if not db:
        return pd.DataFrame(columns=base_cols)
    df = pd.DataFrame(ensure_all_supplier_keys(db, fournisseurs=FOURNISSEURS))
    for col in base_cols:
        if col not in df.columns:
            df[col] = None
    return df[base_cols]


# ──────────────────────────────────────────────────────────────────────────────
# Import Excel
# ──────────────────────────────────────────────────────────────────────────────
def import_from_excel(file):
    raw = pd.read_excel(file, sheet_name=0, header=None)
    rows = []

    for i in range(5, len(raw)):
        row = raw.iloc[i]
        cat = row[1] if len(row) > 1 else None
        desi = row[2] if len(row) > 2 else None
        unit = row[4] if len(row) > 4 else None

        if pd.isna(cat) and pd.isna(desi):
            continue

        d = str(desi).strip() if not pd.isna(desi) else ""
        if not d or d.upper() in ("NAN", "TOTAL HT HORS MO"):
            continue

        prices = {}
        for idx, fn in enumerate(FOURNISSEURS):
            col = 5 + idx
            val = row[col] if col < len(row) else None
            prices[fn] = None if pd.isna(val) else float(val)

        rows.append(
            {
                "id": gen_id(),
                "categorie": str(cat).strip() if not pd.isna(cat) else "Accessoires",
                "designation": d,
                "unite": str(unit).strip() if not pd.isna(unit) else "pcs",
                "note": "",
                **prices,
            }
        )

    return ensure_all_supplier_keys(rows, fournisseurs=FOURNISSEURS)


# ──────────────────────────────────────────────────────────────────────────────
# Exports catalogue
# ──────────────────────────────────────────────────────────────────────────────
def export_catalogue_xlsx(db):
    buf = BytesIO()
    rows = []

    for r in ensure_all_supplier_keys(db, fournisseurs=FOURNISSEURS):
        row = {
            "Catégorie": r.get("categorie", ""),
            "Désignation": r.get("designation", ""),
            "Unité": r.get("unite", ""),
        }
        for fn in FOURNISSEURS:
            row[f"{fn} (FCFA)"] = r.get(fn) or ""
        rows.append(row)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Catalogue")
        ws = writer.sheets["Catalogue"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        navy_f = PatternFill("solid", fgColor="0A2540")
        lgray_f = PatternFill("solid", fgColor="F4F6F8")
        white_f = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)

        widths = [22, 48, 10] + [18] * len(FOURNISSEURS)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for ci in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = navy_f
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd

        for ri in range(2, len(df) + 2):
            fill = lgray_f if ri % 2 == 0 else white_f
            for ci in range(1, len(df.columns) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = fill
                cell.font = Font(size=10)
                cell.border = brd
                cell.alignment = Alignment(vertical="center")
                if ci > 3 and cell.value not in ("", None):
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    except Exception:
                        pass

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf


def export_catalogue_pdf(db):
    supplier_count = len(FOURNISSEURS)
    if supplier_count <= 4:
        pagesize = A4
        left_margin = 12 * mm
        right_margin = 12 * mm
    else:
        pagesize = landscape(A4)
        left_margin = 10 * mm
        right_margin = 10 * mm

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=14 * mm,
        bottomMargin=12 * mm,
    )
    page_width = pagesize[0] - left_margin - right_margin
    stls = getSampleStyleSheet()
    NC = colors.HexColor(NAVY)
    OC = colors.HexColor(ORANGE)
    LG = colors.HexColor("#F4F6F8")

    def ps(n, **kw):
        return ParagraphStyle(n, parent=stls["Normal"], **kw)

    story = []
    hd = [[
        Paragraph(
            "<font color='#F5A623' size='14'><b>INEVOKE SARL</b></font><br/>"
            "<font color='#9ab3c8' size='8'>Catalogue Équipements Solaires</font>",
            ps("hl", textColor=colors.white, leading=18),
        ),
        Paragraph(
            f"<font color='#F5A623' size='11'><b>CATALOGUE PRIX</b></font><br/>"
            f"<font color='#9ab3c8' size='8'>Édité le {date.today().strftime('%d/%m/%Y')}</font>",
            ps("hr", textColor=colors.white, leading=16, alignment=TA_RIGHT),
        ),
    ]]
    ht = Table(hd, colWidths=[page_width * 0.65, page_width * 0.35])
    ht.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), NC),
                ("ROWPADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(ht)
    story.append(Spacer(1, 5 * mm))

    base_widths = [30 * mm, 70 * mm, 16 * mm]
    remaining = page_width - sum(base_widths)
    supplier_width = max(16 * mm, remaining / max(1, supplier_count))
    cw = base_widths + [supplier_width] * supplier_count

    headers = ["Catégorie", "Désignation", "Unité"] + FOURNISSEURS
    ths = ps("th", fontSize=8, textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER)
    tds = ps("td", fontSize=8, leading=11)
    tns = ps("tn", fontSize=8, leading=11, alignment=TA_RIGHT)
    tcs = ps("tc", fontSize=8, leading=11, alignment=TA_CENTER)

    td = [[Paragraph(h, ths) for h in headers]]
    for r in ensure_all_supplier_keys(db, fournisseurs=FOURNISSEURS):
        row_cells = [
            Paragraph(r.get("categorie", ""), tds),
            Paragraph(f"<b>{r.get('designation', '')}</b>", tds),
            Paragraph(r.get("unite", ""), tcs),
        ]
        for fn in FOURNISSEURS:
            row_cells.append(Paragraph(fmt_fcfa(r.get(fn)), tns))
        td.append(row_cells)

    tbl = Table(td, colWidths=cw, repeatRows=1)
    ts = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), NC),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, OC),
            ("ROWPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#e0e0e0")),
        ]
    )
    for i in range(1, len(td)):
        ts.add("BACKGROUND", (0, i), (-1, i), LG if i % 2 == 0 else colors.white)

    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=NC))
    story.append(Spacer(1, 2 * mm))
    story.append(
        Paragraph(
            f"<font size='7' color='#888'>INEVOKE SARL · Abidjan, Côte d'Ivoire · (225) 0544125825 · "
            f"hello@inevoke.ci · www.inevoke.ci · {date.today().strftime('%d/%m/%Y')}</font>",
            ps("ft", alignment=TA_CENTER),
        )
    )
    doc.build(story)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# Génération devis PDF
# ──────────────────────────────────────────────────────────────────────────────
def generate_devis_pdf(info, lignes, fournisseur, remise_pct, tva_pct):
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    stls = getSampleStyleSheet()
    NC = colors.HexColor(NAVY)
    OC = colors.HexColor(ORANGE)
    LG = colors.HexColor("#F4F6F8")

    def ps(n, **kw):
        return ParagraphStyle(n, parent=stls["Normal"], **kw)

    story = []
    hd = [[
        Paragraph(
            "<font color='#F5A623'><b>INEVOKE SARL</b></font><br/>"
            "<font size='8' color='#9ab3c8'>Énergies Renouvelables · Facility Management · Digital</font><br/>"
            "<font size='8' color='#9ab3c8'>Abidjan, Côte d'Ivoire · contact@inevoke.ci</font>",
            ps("hl", fontSize=12, textColor=colors.white, leading=18),
        ),
        Paragraph(
            f"<b><font size='18' color='#F5A623'>DEVIS</font></b><br/>"
            f"<font size='9' color='#9ab3c8'>N° {info['numero']}</font><br/>"
            f"<font size='8' color='#9ab3c8'>Date : {info['date']}</font><br/>"
            f"<font size='8' color='#9ab3c8'>Validité : {info['validite']}</font>",
            ps("hr", fontSize=9, textColor=colors.white, leading=16, alignment=TA_RIGHT),
        ),
    ]]
    ht = Table(hd, colWidths=[110 * mm, 65 * mm])
    ht.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), NC),
                ("ROWPADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(ht)
    story.append(Spacer(1, 5 * mm))

    inf = [[
        Paragraph(
            f"<b>CLIENT</b><br/><font size='10'>{info['client']}</font><br/>"
            f"<font size='9' color='#555'>{info['contact']}</font><br/>"
            f"<font size='9' color='#555'>{info['adresse']}</font>",
            ps("cl", leading=15),
        ),
        Paragraph(
            f"<b>PROJET</b><br/><font size='10'>{info['projet']}</font><br/>"
            f"<font size='9' color='#555'>Fournisseur : <b>{fournisseur}</b></font><br/>"
            f"<font size='9' color='#555'>{info['notes']}</font>",
            ps("pr", leading=15),
        ),
    ]]
    it = Table(inf, colWidths=[87 * mm, 88 * mm])
    it.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), LG),
                ("ROWPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEAFTER", (0, 0), (0, -1), 0.5, colors.HexColor("#ddd")),
            ]
        )
    )
    story.append(it)
    story.append(Spacer(1, 5 * mm))

    cw = [8 * mm, 26 * mm, 62 * mm, 10 * mm, 13 * mm, 26 * mm, 26 * mm]
    hr2 = ["N°", "Catégorie", "Désignation", "Qté", "Unité", "P.U. (FCFA)", "Total HT"]

    ths2 = ps("th2", fontSize=9, textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER)
    tds2 = ps("td2", fontSize=9, leading=12)
    tns2 = ps("tn2", fontSize=9, leading=12, alignment=TA_RIGHT)
    tcs2 = ps("tc2", fontSize=9, leading=12, alignment=TA_CENTER)
    cts2 = ps("ct2", fontSize=8, textColor=colors.HexColor("#666"), leading=11)

    tdata = [[Paragraph(h, ths2) for h in hr2]]
    total_ht = 0

    for idx, lg in enumerate(lignes, 1):
        pu = lg.get(fournisseur) or 0
        qt = float(lg.get("quantite", 1))
        tot = pu * qt
        total_ht += tot

        tdata.append(
            [
                Paragraph(str(idx), tcs2),
                Paragraph(lg.get("categorie", ""), cts2),
                Paragraph(f"<b>{lg.get('designation', '')}</b>", tds2),
                Paragraph(str(int(qt)), tcs2),
                Paragraph(lg.get("unite", ""), tcs2),
                Paragraph(f"{pu:,.0f}".replace(",", "  "), tns2),
                Paragraph(f"{tot:,.0f}".replace(",", "  "), tns2),
            ]
        )

    mt = Table(tdata, colWidths=cw, repeatRows=1)
    ms = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), NC),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, OC),
            ("ROWPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#e0e0e0")),
        ]
    )
    for i in range(1, len(tdata)):
        ms.add("BACKGROUND", (0, i), (-1, i), LG if i % 2 == 0 else colors.white)

    mt.setStyle(ms)
    story.append(mt)
    story.append(Spacer(1, 4 * mm))

    rv = total_ht * (remise_pct / 100)
    bv = total_ht - rv
    tv = bv * (tva_pct / 100)
    ttc = bv + tv

    tots = [
        ["", "", "Total HT", f"{total_ht:,.0f} FCFA"],
        ["", "", f"Remise ({remise_pct:.0f}%)", f"- {rv:,.0f} FCFA"],
        ["", "", f"TVA ({tva_pct:.0f}%)", f"{tv:,.0f} FCFA"],
        ["", "", "TOTAL TTC", f"{ttc:,.0f} FCFA"],
    ]
    tt = Table(tots, colWidths=[60 * mm, 30 * mm, 50 * mm, 35 * mm])
    tt.setStyle(
        TableStyle(
            [
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ROWPADDING", (0, 0), (-1, -1), 4),
                ("LINEABOVE", (2, -1), (-1, -1), 1, NC),
                ("FONTNAME", (2, -1), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (2, -1), (-1, -1), 13),
                ("TEXTCOLOR", (2, -1), (-1, -1), NC),
                ("BACKGROUND", (2, -1), (-1, -1), colors.HexColor("#FFF3E0")),
                ("LINEBELOW", (2, 1), (-1, 1), 0.4, colors.HexColor("#ccc")),
            ]
        )
    )
    story.append(tt)
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#ddd")))
    story.append(Spacer(1, 3 * mm))

    cond = [[
        Paragraph(
            "<b>Conditions générales</b><br/>"
            "<font size='8'>• Devis valable 30 jours à compter de la date d'émission.<br/>"
            "• Paiement : 50% à la commande, 50% à la réception.<br/>"
            "• Délai de livraison : selon disponibilité (5-15 jours).<br/>"
            "• Garantie : selon garantie constructeur.</font>",
            ps("cd", leading=13),
        ),
        Paragraph(
            "<b>Cachet &amp; Signature</b><br/><br/>"
            "<font size='8' color='#555'>INEVOKE SARL — Direction Générale</font>"
            "<br/><br/><br/><font size='8'>___________________________</font>",
            ps("sg", leading=13, alignment=TA_RIGHT),
        ),
    ]]
    ct = Table(cond, colWidths=[100 * mm, 75 * mm])
    ct.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWPADDING", (0, 0), (-1, -1), 4)]))
    story.append(ct)
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=NC))
    story.append(Spacer(1, 2 * mm))
    story.append(
        Paragraph(
            "<font size='7' color='#888'>INEVOKE SARL — RCCM CI-ABJ-2024 — "
            "Abidjan, Côte d'Ivoire — contact@inevoke.ci — www.inevoke.ci</font>",
            ps("ft", alignment=TA_CENTER),
        )
    )

    doc.build(story)
    buf.seek(0)
    return buf, ttc


# ──────────────────────────────────────────────────────────────────────────────
# Helpers devis
# ──────────────────────────────────────────────────────────────────────────────
def reset_generated_quote():
    st.session_state.devis_pdf_bytes = None
    st.session_state.devis_pdf_name = None
    st.session_state.devis_total_ttc = None


def add_line_to_quote(line_dict):
    existing = next(
        (l for l in st.session_state.lignes_devis if l["id"] == line_dict["id"]),
        None,
    )
    if existing:
        existing["quantite"] += 1
    else:
        lg = deepcopy(line_dict)
        lg["quantite"] = 1
        st.session_state.lignes_devis.append(lg)
    reset_generated_quote()


def remove_line_from_quote(index_to_remove):
    if 0 <= index_to_remove < len(st.session_state.lignes_devis):
        st.session_state.lignes_devis.pop(index_to_remove)
    reset_generated_quote()


def clear_quote():
    st.session_state.lignes_devis = []
    reset_generated_quote()


# ──────────────────────────────────────────────────────────────────────────────
# INTERFACE
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
<div class="main-header">
  <div style="font-size:38px">☀️</div>
  <div>
    <h1>INEVOKE SARL — Outil de Devis Solaire</h1>
    <p>Catalogue équipements · Création de devis professionnels · Export PDF &amp; Excel</p>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown(
        f"<div style='font-size:13px;font-weight:600;color:{NAVY};padding:.5rem 0'>📋 Navigation</div>",
        unsafe_allow_html=True,
    )
    page = st.radio(
        "",
        ["📦 Catalogue", "➕ Ajouter un équipement", "✏️ Modifier le catalogue", "📄 Créer un devis"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown(
        "<div style='font-size:11px;color:#888;text-align:center'>INEVOKE SARL © 2026<br/>v2.0</div>",
        unsafe_allow_html=True,
    )

show_flash_messages()
db = load_db()


# ──────────────────────────────────────────────────────────────────────────────
# PAGE 1 — CATALOGUE
# ──────────────────────────────────────────────────────────────────────────────
if page == "📦 Catalogue":
    st.markdown("<div class='section-title'>📦 Catalogue équipements</div>", unsafe_allow_html=True)
    df_db = db_to_df(db)

    st.markdown("<br>", unsafe_allow_html=True)

    with st.expander("📥 Importer depuis un fichier Excel"):
        st.markdown(
            "<div class='info-box'>Importez votre fichier Excel. "
            "Les colonnes fournisseurs seront lues selon la liste actuelle des fournisseurs.</div>",
            unsafe_allow_html=True,
        )
        up_file = st.file_uploader("Choisir le fichier", type=["xlsx"], key="up1")
        if up_file:
            rows = import_from_excel(up_file)
            st.info(f"📋 {len(rows)} équipements détectés.")
            mode = st.radio(
                "Mode d'import",
                [
                    "Fusionner (ajouter les nouveaux, mettre à jour les existants)",
                    "Remplacer intégralement la base",
                ],
                key="import_mode",
            )
            if st.button("✅ Confirmer l'import", key="btn_import_confirm"):
                if "Remplacer" in mode:
                    db = rows
                else:
                    desigs = {r["designation"].lower(): i for i, r in enumerate(db)}
                    for r in rows:
                        k = r["designation"].lower()
                        if k not in desigs:
                            db.append(r)
                            desigs[k] = len(db) - 1
                        else:
                            existing = db[desigs[k]]
                            for fn in FOURNISSEURS:
                                if not existing.get(fn) and r.get(fn):
                                    existing[fn] = r[fn]
                save_db(db)
                st.success("✅ Import terminé !")

    cf1, cf2, cf3 = st.columns([1, 1, 2])
    with cf1:
        cat_f = st.selectbox("Catégorie", ["Toutes"] + CATEGORIES, key="cf1")
    with cf2:
        fn_f = st.selectbox("Fournisseur", ["Tous"] + FOURNISSEURS, key="cf2")
    with cf3:
        srch = st.text_input("🔍 Recherche", key="srch1", placeholder="ex: panneau, câble...")

    df_v = df_db.copy()
    if not df_v.empty:
        if cat_f != "Toutes":
            df_v = df_v[df_v["categorie"] == cat_f]
        if fn_f != "Tous":
            df_v = df_v[df_v[fn_f].notna() & (df_v[fn_f] > 0)]
        if srch:
            df_v = df_v[df_v["designation"].str.contains(srch, case=False, na=False)]

    nb_equipements = len(df_v)
    nb_categories = df_v["categorie"].nunique() if not df_v.empty else 0
    if not df_v.empty:
        nb_fournisseurs = sum(
            1
            for fn in FOURNISSEURS
            if fn in df_v.columns and df_v[fn].notna().any() and (df_v[fn] > 0).any()
        )
    else:
        nb_fournisseurs = 0

    c1, c2, c3, c4 = st.columns(4)
    kpi_data = [
        (c1, nb_equipements, "Équipements", ""),
        (c2, nb_categories, "Catégories", ""),
        (c3, nb_fournisseurs, "Fournisseurs actifs", ""),
        (c4, "FCFA", "Devise", f"color:{ORANGE}"),
    ]
    for col, val, lbl, style in kpi_data:
        with col:
            st.markdown(
                f"<div class='kpi-card'>"
                f"<div class='kpi-val' style='{style}'>{val}</div>"
                f"<div class='kpi-lbl'>{lbl}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    if not df_v.empty:
        df_show = df_v[["categorie", "designation", "unite"] + FOURNISSEURS].copy()
        df_show.columns = ["Catégorie", "Désignation", "Unité"] + [f"{fn}\n(FCFA)" for fn in FOURNISSEURS]
        for fn in FOURNISSEURS:
            col_name = f"{fn}\n(FCFA)"
            df_show[col_name] = df_show[col_name].apply(lambda x: fmt_fcfa(x))
        st.dataframe(df_show, use_container_width=True, hide_index=True)
        st.caption(f"{len(df_v)} équipement(s) affiché(s)")
    else:
        st.markdown(
            "<div class='warn-box'>📭 Aucun équipement correspondant aux filtres. "
            "Modifiez les filtres ou importez des équipements via <b>Ajouter un équipement</b>.</div>",
            unsafe_allow_html=True,
        )

    if db:
        st.markdown(
            "<div class='section-title'>📥 Télécharger le catalogue complet</div>",
            unsafe_allow_html=True,
        )
        d1, d2 = st.columns(2)
        with d1:
            xls = export_catalogue_xlsx(db)
            st.download_button(
                "📊 Télécharger en Excel (.xlsx)",
                data=xls,
                file_name=f"INEVOKE_Catalogue_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_catalog_xlsx",
            )
        with d2:
            pdf_c = export_catalogue_pdf(db)
            st.download_button(
                "📄 Télécharger en PDF",
                data=pdf_c,
                file_name=f"INEVOKE_Catalogue_{date.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="download_catalog_pdf",
            )


# ──────────────────────────────────────────────────────────────────────────────
# PAGE 2 — AJOUTER UN ÉQUIPEMENT
# ──────────────────────────────────────────────────────────────────────────────
elif page == "➕ Ajouter un équipement":
    st.markdown("<div class='section-title'>➕ Ajouter un équipement</div>", unsafe_allow_html=True)

    cc1, cc2 = st.columns([2, 1])
    with cc1:
        cat_ch = st.selectbox(
            "Catégorie *",
            ["— Choisir —"] + CATEGORIES + ["+ Nouvelle catégorie..."],
            key="add_single_cat_select",
        )
    with cc2:
        new_cat = ""
        if cat_ch == "+ Nouvelle catégorie...":
            new_cat = st.text_input(
                "Nom de la nouvelle catégorie",
                placeholder="ex: Pompes solaires",
                key="add_single_new_cat",
            )

    cat_final = new_cat.strip() if cat_ch == "+ Nouvelle catégorie..." else cat_ch

    st.markdown("**Équipement unique**")
    with st.form("form_add", clear_on_submit=True):
        cd1, cd2 = st.columns([3, 1])
        with cd1:
            desi = st.text_input(
                "Désignation *",
                placeholder="ex: Panneau solaire 400W monocristallin",
                key="add_single_designation",
            )
        with cd2:
            unite = st.text_input("Unité *", value="pcs", placeholder="pcs / lot / m...", key="add_single_unite")

        st.markdown("**Prix unitaires (FCFA) — Laisser 0 si non disponible**")
        fc = st.columns(max(1, len(FOURNISSEURS)))
        prix = {}
        for i, fn in enumerate(FOURNISSEURS):
            with fc[i]:
                v = st.number_input(
                    fn,
                    min_value=0.0,
                    step=100.0,
                    key=f"ap_{fn}",
                    format="%.0f",
                )
                prix[fn] = float(v) if v > 0 else None

        note = st.text_input(
            "Note (optionnel)",
            placeholder="ex: Compatible 48V — garantie 10 ans",
            key="add_single_note",
        )
        ok = st.form_submit_button("💾 Enregistrer", type="primary", use_container_width=True)

        if ok:
            if not desi.strip():
                st.error("⚠️ Désignation obligatoire.")
            elif cat_final in ("— Choisir —", ""):
                st.error("⚠️ Choisissez ou créez une catégorie.")
            elif any(r["designation"].lower() == desi.strip().lower() for r in db):
                st.warning(f"⚠️ « {desi} » existe déjà. Utilisez **Modifier le catalogue** pour le mettre à jour.")
            else:
                if cat_final not in CATEGORIES:
                    CATEGORIES.append(cat_final)
                    persist_categories(CATEGORIES)

                db.append(
                    {
                        "id": gen_id(),
                        "categorie": cat_final,
                        "designation": desi.strip(),
                        "unite": unite.strip() or "pcs",
                        "note": note.strip(),
                        **prix,
                    }
                )
                save_db(db)
                st.success(f"✅ « {desi} » ajouté dans {cat_final} !")

    st.markdown("---")
    st.markdown(
        "<div class='section-title'>📋 Ajout en masse — tableau interactif</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='info-box'>Remplissez le tableau ci-dessous ou collez vos lignes "
        "depuis Excel. Ajoutez autant de lignes que nécessaire avec le bouton + en bas.<br/>"
        "<b>Astuce :</b> tapez une nouvelle catégorie directement dans la colonne Catégorie "
        "— elle sera automatiquement ajoutée à la liste déroulante du catalogue.</div>",
        unsafe_allow_html=True,
    )

    tmpl_data = []
    for _ in range(3):
        row = {"Catégorie": "", "Désignation": "", "Unité": "pcs"}
        for fn in FOURNISSEURS:
            row[fn] = 0
        tmpl_data.append(row)
    tmpl = pd.DataFrame(tmpl_data)

    bulk_config = {
        "Catégorie": st.column_config.TextColumn(
            "Catégorie",
            help="Tapez une catégorie existante ou créez-en une nouvelle. "
            "Elle sera ajoutée automatiquement à la liste du catalogue.",
            required=False,
        ),
        "Désignation": st.column_config.TextColumn("Désignation", required=False),
        "Unité": st.column_config.TextColumn("Unité"),
    }
    for fn in FOURNISSEURS:
        bulk_config[fn] = st.column_config.NumberColumn(format="%.0f", min_value=0)

    st.caption(f"Catégories existantes : {', '.join(CATEGORIES)}")

    edited = st.data_editor(
        tmpl,
        num_rows="dynamic",
        use_container_width=True,
        column_config=bulk_config,
        key="bulk_ed",
    )

    if st.button("💾 Enregistrer le tableau", use_container_width=True, key="bulk_save_btn"):
        desigs = {r["designation"].lower() for r in db}
        added = 0
        skipped = 0
        nouvelles_cats = 0

        for _, row in edited.iterrows():
            d = str(row.get("Désignation", "")).strip()
            c = str(row.get("Catégorie", "")).strip()

            if not d or not c:
                continue

            if d.lower() in desigs:
                skipped += 1
                continue

            if c not in CATEGORIES:
                CATEGORIES.append(c)
                nouvelles_cats += 1

            item = {
                "id": gen_id(),
                "categorie": c,
                "designation": d,
                "unite": str(row.get("Unité", "pcs")).strip() or "pcs",
                "note": "",
            }
            for fn in FOURNISSEURS:
                item[fn] = safe_float(row.get(fn))

            db.append(item)
            desigs.add(d.lower())
            added += 1

        if nouvelles_cats:
            persist_categories(CATEGORIES)

        save_db(db)

        msg = f"✅ {added} équipement(s) ajouté(s)."
        if skipped:
            msg += f" {skipped} ignoré(s) (doublon)."
        if nouvelles_cats:
            msg += f" {nouvelles_cats} nouvelle(s) catégorie(s) créée(s)."
        st.success(msg)


# ──────────────────────────────────────────────────────────────────────────────
# PAGE 3 — MODIFIER LE CATALOGUE
# ──────────────────────────────────────────────────────────────────────────────
elif page == "✏️ Modifier le catalogue":
    st.markdown(
        "<div class='section-title'>✏️ Modifier le catalogue — Édition complète</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='info-box'><b>Mode édition libre</b> — "
        "Ajoutez des lignes en bas du tableau, modifiez les cellules, "
        "puis cliquez sur <b>Enregistrer le tableau</b>.</div>",
        unsafe_allow_html=True,
    )

    df_db = db_to_df(db)
    if df_db.empty:
        st.warning("Aucun équipement dans la base.")
        st.stop()

    st.markdown("<div class='section-title'>🧱 Gérer les colonnes fournisseurs</div>", unsafe_allow_html=True)

    g1, g2 = st.columns([2, 1])
    with g1:
        new_supplier = st.text_input(
            "Ajouter un nouveau fournisseur",
            placeholder="ex: SMA",
            key="new_supplier_name",
        )
    with g2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ Ajouter la colonne", use_container_width=True, key="btn_add_supplier"):
            nom = new_supplier.strip().upper()
            if not nom:
                st.warning("⚠️ Entrez un nom de fournisseur.")
            elif nom in FOURNISSEURS:
                st.warning(f"⚠️ Le fournisseur « {nom} » existe déjà.")
            else:
                FOURNISSEURS.append(nom)
                persist_fournisseurs(FOURNISSEURS)

                for r in db:
                    if nom not in r:
                        r[nom] = None
                save_db(db)

                st.success(f"✅ Colonne fournisseur « {nom} » ajoutée.")
                st.stop()

    g3, g4 = st.columns([2, 1])
    with g3:
        supplier_to_delete = st.selectbox(
            "Supprimer un fournisseur",
            ["— Choisir —"] + FOURNISSEURS,
            key="supplier_to_delete",
        )
    with g4:
        st.markdown("<br>", unsafe_allow_html=True)
        if supplier_to_delete != "— Choisir —" and st.button(
            "🗑️ Supprimer la colonne",
            use_container_width=True,
            key="btn_delete_supplier",
        ):
            FOURNISSEURS = [f for f in FOURNISSEURS if f != supplier_to_delete]
            persist_fournisseurs(FOURNISSEURS)

            for r in db:
                if supplier_to_delete in r:
                    del r[supplier_to_delete]
            save_db(db)

            st.success(f"✅ Colonne « {supplier_to_delete} » supprimée.")
            st.stop()

    st.markdown("---")

    fe1, fe2 = st.columns([1, 2])
    with fe1:
        cat_e = st.selectbox("Filtrer catégorie", ["Toutes"] + CATEGORIES, key="edit_cat_filter")
    with fe2:
        srch_e = st.text_input("🔍 Rechercher", key="edit_search_filter", placeholder="Filtrer...")

    df_e = df_db.copy()
    if cat_e != "Toutes":
        df_e = df_e[df_e["categorie"] == cat_e]
    if srch_e:
        df_e = df_e[df_e["designation"].str.contains(srch_e, case=False, na=False)]

    st.caption(f"{len(df_e)} équipement(s) affiché(s) — ajout de lignes possible en bas du tableau")

    base_cols = ["id", "categorie", "designation", "unite"]
    for fn in FOURNISSEURS:
        if fn not in df_e.columns:
            df_e[fn] = None
    df_ed = df_e[base_cols + FOURNISSEURS].copy()

    column_config = {
        "id": st.column_config.TextColumn("ID", disabled=True),
        "categorie": st.column_config.SelectboxColumn("Catégorie", options=CATEGORIES, required=True),
        "designation": st.column_config.TextColumn("Désignation", width="large", required=True),
        "unite": st.column_config.TextColumn("Unité", width="small"),
    }
    for fn in FOURNISSEURS:
        column_config[fn] = st.column_config.NumberColumn(f"{fn} (FCFA)", format="%.0f", min_value=0)

    edited_df = st.data_editor(
        df_ed,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config=column_config,
        key="main_editor_table",
    )

    if st.button("💾 Enregistrer le tableau", type="primary", use_container_width=True, key="save_catalog_table"):
        filtered_ids = set(df_e["id"].dropna().tolist()) if "id" in df_e.columns else set()
        orig_map = {r["id"]: deepcopy(r) for r in db if r.get("id")}
        new_db = []

        for r in db:
            if r.get("id") not in filtered_ids:
                nr = normalize_record(r, fournisseurs=FOURNISSEURS)
                for fn in FOURNISSEURS:
                    if fn not in nr:
                        nr[fn] = None
                new_db.append(nr)

        for _, row in edited_df.iterrows():
            designation = str(row.get("designation", "")).strip()
            categorie = str(row.get("categorie", "")).strip()
            unite = str(row.get("unite", "pcs")).strip() or "pcs"

            if not designation or not categorie:
                continue

            row_id = row.get("id")
            if pd.isna(row_id) or not str(row_id).strip():
                row_id = gen_id()

            old_row = orig_map.get(row_id, {})
            new_row = {
                "id": row_id,
                "categorie": categorie,
                "designation": designation,
                "unite": unite,
                "note": old_row.get("note", ""),
            }

            for fn in FOURNISSEURS:
                v = row.get(fn)
                try:
                    new_row[fn] = float(v) if pd.notna(v) and float(v) > 0 else None
                except Exception:
                    new_row[fn] = None

            new_db.append(new_row)

            if categorie not in CATEGORIES:
                CATEGORIES.append(categorie)

        persist_categories(CATEGORIES)

        final = []
        seen = set()
        for r in new_db:
            k = r["designation"].strip().lower()
            if k not in seen:
                seen.add(k)
                final.append(normalize_record(r, fournisseurs=FOURNISSEURS))

        save_db(final)
        st.success(f"✅ {len(final)} équipement(s) enregistrés avec succès.")

    s2, s3 = st.columns([1, 1])
    with s2:
        eq_del = st.selectbox(
            "Supprimer un équipement",
            ["— Choisir —"] + df_e["designation"].tolist(),
            key="del2_select",
        )
    with s3:
        st.markdown("<br>", unsafe_allow_html=True)
        if eq_del != "— Choisir —" and st.button(
            "🗑️ Supprimer l'équipement",
            use_container_width=True,
            key="del2_btn",
        ):
            db = [r for r in db if r["designation"] != eq_del]
            save_db(db)
            st.success(f"🗑️ « {eq_del} » supprimé.")
            st.stop()

    st.markdown(
        "<div class='section-title'>🔍 Fiche individuelle — modifier un équipement précis</div>",
        unsafe_allow_html=True,
    )
    eq_ch = st.selectbox(" — ", [" — "] + [r["designation"] for r in db], key="fiche2")
    if eq_ch != " — ":
        eq = next((r for r in db if r["designation"] == eq_ch), None)
        if eq:
            with st.form("fiche_form"):
                fa, fb, fc = st.columns([2, 1, 1])
                with fa:
                    nd = st.text_input("Désignation", value=eq.get("designation", ""), key="fiche_designation")
                with fb:
                    ci = CATEGORIES.index(eq.get("categorie", "Accessoires")) if eq.get("categorie") in CATEGORIES else 0
                    nc = st.selectbox("Catégorie", CATEGORIES, index=ci, key="fiche_categorie")
                with fc:
                    nu = st.text_input("Unité", value=eq.get("unite", "pcs"), key="fiche_unite")

                st.markdown("**Prix par fournisseur (FCFA)**")
                fc2 = st.columns(max(1, len(FOURNISSEURS)))
                np2 = {}
                for i, fn in enumerate(FOURNISSEURS):
                    with fc2[i]:
                        v = st.number_input(
                            fn,
                            value=float(eq.get(fn) or 0),
                            min_value=0.0,
                            step=100.0,
                            format="%.0f",
                            key=f"fp_{fn}",
                        )
                        np2[fn] = float(v) if v > 0 else None

                nn = st.text_input("Note", value=eq.get("note", ""), key="fiche_note")

                if st.form_submit_button("💾 Mettre à jour cette fiche", use_container_width=True):
                    for r in db:
                        if r["designation"] == eq_ch:
                            r.update(
                                {
                                    "designation": nd.strip(),
                                    "categorie": nc,
                                    "unite": nu.strip() or "pcs",
                                    "note": nn.strip(),
                                    **np2,
                                }
                            )
                    if nc not in CATEGORIES:
                        CATEGORIES.append(nc)
                        persist_categories(CATEGORIES)

                    save_db(db)
                    st.success(f"✅ « {nd} » mis à jour !")


# ──────────────────────────────────────────────────────────────────────────────
# PAGE 4 — CRÉER UN DEVIS
# ──────────────────────────────────────────────────────────────────────────────
elif page == "📄 Créer un devis":
    st.markdown("<div class='section-title'>📄 Créer un devis</div>", unsafe_allow_html=True)
    df_db = db_to_df(db)

    if df_db.empty:
        st.warning("⚠️ Catalogue vide. Allez dans **Catalogue** pour importer vos équipements.")
        st.stop()

    with st.expander("👤 Client & Projet", expanded=True):
        cv1, cv2 = st.columns(2)
        with cv1:
            client = st.text_input("Nom du client *", placeholder="ex: Hôtel Azalaï Abidjan", key="dv_client")
            contact = st.text_input(
                "Contact",
                placeholder="ex: M. Kouassi — +225 07 XX XX XX",
                key="dv_contact",
            )
            adresse = st.text_input("Adresse", placeholder="ex: Plateau, Abidjan", key="dv_adresse")
        with cv2:
            projet = st.text_input(
                "Intitulé du projet *",
                placeholder="ex: Installation solaire 5kW",
                key="dv_projet",
            )
            num_dev = st.text_input(
                "N° Devis",
                value=f"DEV-{date.today().strftime('%Y%m%d')}-{len(db)%100:02d}",
                key="dv_numero",
            )
            date_val = st.date_input("Validité", value=date.today() + timedelta(days=30), key="dv_validite")
            notes = st.text_area("Notes", height=65, placeholder="ex: Site off-grid", key="dv_notes")

    with st.expander("💰 Fournisseur & Conditions", expanded=True):
        cv3, cv4, cv5 = st.columns(3)
        with cv3:
            fournisseur = st.selectbox("Fournisseur retenu", FOURNISSEURS, key="dv_fournisseur")
        with cv4:
            remise = st.number_input("Remise (%)", 0.0, 50.0, 0.0, 0.5, key="dv_remise")
        with cv5:
            tva = st.number_input("TVA (%)", 0.0, 25.0, 18.0, 0.5, key="dv_tva")

    st.markdown("<div class='section-title'>🛒 Sélectionner les équipements</div>", unsafe_allow_html=True)

    fs1, fs2 = st.columns([1, 2])
    with fs1:
        cat_dv = st.selectbox("Catégorie", ["Toutes"] + CATEGORIES, key="cdv")
    with fs2:
        srch_dv = st.text_input("🔍 Rechercher", key="sdv", placeholder="panneau, câble...")

    df_f = df_db.copy()
    if cat_dv != "Toutes":
        df_f = df_f[df_f["categorie"] == cat_dv]
    if srch_dv:
        df_f = df_f[df_f["designation"].str.contains(srch_dv, case=False, na=False)]

    if not df_f.empty:
        for _, row in df_f.iterrows():
            pu = row.get(fournisseur)
            pu_s = f"{pu:,.0f} FCFA".replace(",", "  ") if pd.notna(pu) and pu else "Prix N/D"

            ceq, cpu, cbt = st.columns([4, 2, 1])
            with ceq:
                st.markdown(
                    f"<span style='font-size:11px;color:#777'>{row['categorie']}</span><br>"
                    f"<b style='font-size:13px'>{row['designation']}</b>",
                    unsafe_allow_html=True,
                )
            with cpu:
                clr = NAVY if pd.notna(pu) and pu else "#aaa"
                st.markdown(
                    f"<div style='padding-top:14px;font-size:13px;color:{clr};font-weight:600'>{pu_s}</div>",
                    unsafe_allow_html=True,
                )
            with cbt:
                if st.button("➕", key=f"add_{row['id']}"):
                    add_line_to_quote(row.to_dict())
            st.markdown("<hr style='margin:3px 0;border-color:#f0f0f0'>", unsafe_allow_html=True)

    st.markdown("<div class='section-title'>🧾 Lignes du devis</div>", unsafe_allow_html=True)

    if not st.session_state.lignes_devis:
        st.info("Aucun équipement sélectionné. Cliquez ➕ ci-dessus.")
    else:
        total_ht = 0

        for i, lg in enumerate(st.session_state.lignes_devis):
            pu = lg.get(fournisseur) or 0

            la, lb, lc, ld, le, lf = st.columns([3.5, 0.8, 0.7, 1.2, 1.4, 0.4])
            with la:
                st.markdown(
                    f"<div style='padding-top:8px;font-size:13px'><b>{lg['designation']}</b> "
                    f"<span style='color:#888;font-size:11px'>({lg['categorie']})</span></div>",
                    unsafe_allow_html=True,
                )
            with lb:
                qty = st.number_input(
                    "",
                    value=int(lg["quantite"]),
                    min_value=1,
                    step=1,
                    key=f"q_{lg['id']}",
                    label_visibility="collapsed",
                )
                st.session_state.lignes_devis[i]["quantite"] = qty
            with lc:
                st.markdown(
                    f"<div style='padding-top:8px;font-size:12px;color:#666'>{lg.get('unite', 'pcs')}</div>",
                    unsafe_allow_html=True,
                )
            with ld:
                st.markdown(
                    f"<div style='padding-top:8px;font-size:12px;color:#555'>{pu:,.0f}/u</div>".replace(",", "  "),
                    unsafe_allow_html=True,
                )
            with le:
                tot = pu * qty
                total_ht += tot
                st.markdown(
                    f"<div style='padding-top:8px;font-size:13px;font-weight:700;color:{NAVY}'>"
                    f"{tot:,.0f} FCFA</div>".replace(",", "  "),
                    unsafe_allow_html=True,
                )
            with lf:
                if st.button("🗑", key=f"rm_{lg['id']}"):
                    remove_line_from_quote(i)
                    st.stop()

            st.markdown("<hr style='margin:2px 0;border-color:#f0f0f0'>", unsafe_allow_html=True)

        with st.expander("➕ Ajouter une ligne libre (hors catalogue)"):
            hc1, hc2, hc3, hc4 = st.columns([3, 1, 1, 1])
            with hc1:
                hd = st.text_input("Désignation", key="hd", placeholder="ex: Transport")
            with hc2:
                hq = st.number_input("Qté", min_value=1, value=1, key="hq")
            with hc3:
                hp = st.number_input("P.U. FCFA", min_value=0.0, value=0.0, key="hp", step=500.0)
            with hc4:
                hu = st.text_input("Unité", value="lot", key="hu")

            if st.button("➕ Ajouter", key="hadd"):
                if hd.strip():
                    lg2 = {
                        "id": gen_id(),
                        "categorie": "Hors catalogue",
                        "designation": hd.strip(),
                        "unite": hu.strip() or "lot",
                        "quantite": int(hq),
                    }
                    for fn in FOURNISSEURS:
                        lg2[fn] = float(hp)
                    st.session_state.lignes_devis.append(lg2)
                    reset_generated_quote()
                    st.success("✅ Ligne libre ajoutée.")

        st.markdown("---")
        rv = total_ht * (remise / 100)
        bv = total_ht - rv
        tv = bv * (tva / 100)
        ttc = bv + tv

        _, rc = st.columns([2, 1])
        with rc:
            st.markdown(
                f"""
            <div style='background:#F4F6F8;border-radius:10px;padding:16px'>
              <div style='display:flex;justify-content:space-between;font-size:13px;margin-bottom:5px'>
                <span>Total HT</span><b>{total_ht:,.0f} FCFA</b></div>
              <div style='display:flex;justify-content:space-between;font-size:13px;margin-bottom:5px;color:#666'>
                <span>Remise ({remise:.0f}%)</span><span>- {rv:,.0f} FCFA</span></div>
              <div style='display:flex;justify-content:space-between;font-size:13px;margin-bottom:5px;color:#666'>
                <span>TVA ({tva:.0f}%)</span><span>{tv:,.0f} FCFA</span></div>
              <hr style='border-color:#ddd;margin:8px 0'/>
              <div style='display:flex;justify-content:space-between;font-size:17px;font-weight:700;color:{NAVY}'>
                <span>TOTAL TTC</span><span style='color:{ORANGE}'>{ttc:,.0f} FCFA</span></div>
            </div>
            """.replace(",", "  "),
                unsafe_allow_html=True,
            )

        pb1, pb2 = st.columns([3, 1])
        with pb2:
            if st.button("🗑️ Vider le devis", use_container_width=True, key="clear_quote_btn"):
                clear_quote()
                st.success("✅ Devis vidé.")
                st.stop()

        st.markdown("<br>", unsafe_allow_html=True)

        if st.button("📥 Générer le devis PDF", type="primary", use_container_width=True, key="generate_quote_pdf"):
            if not client.strip() or not projet.strip():
                st.error("⚠️ Renseignez le nom du client et l'intitulé du projet.")
            else:
                devis_info = {
                    "numero": num_dev,
                    "date": date.today().strftime("%d/%m/%Y"),
                    "validite": date_val.strftime("%d/%m/%Y"),
                    "client": client,
                    "contact": contact or "—",
                    "adresse": adresse or "—",
                    "projet": projet,
                    "notes": notes or "—",
                }
                pdf_buf, total_ttc = generate_devis_pdf(
                    devis_info,
                    st.session_state.lignes_devis,
                    fournisseur,
                    remise,
                    tva,
                )

                nom = f"Devis_{client.replace(' ', '_')}_{num_dev}.pdf"
                st.session_state.devis_pdf_bytes = pdf_buf.getvalue()
                st.session_state.devis_pdf_name = nom
                st.session_state.devis_total_ttc = total_ttc
                st.success("✅ PDF généré. Vous pouvez maintenant le télécharger ci-dessous.")

        if st.session_state.devis_pdf_bytes:
            st.download_button(
                label=f"⬇️ Télécharger — {st.session_state.devis_pdf_name}",
                data=st.session_state.devis_pdf_bytes,
                file_name=st.session_state.devis_pdf_name,
                mime="application/pdf",
                use_container_width=True,
                key="download_generated_quote_pdf",
            )
            st.markdown(
                f"""<div class='success-box'>✅ Devis <b>{num_dev}</b> généré pour 
                <b>{client}</b> · Fournisseur : <b>{fournisseur}</b> · 
                {len(st.session_state.lignes_devis)} ligne(s) · 
                Total TTC : <b>{st.session_state.devis_total_ttc:,.0f} FCFA</b></div>""".replace(",", "  "),
                unsafe_allow_html=True,
            )